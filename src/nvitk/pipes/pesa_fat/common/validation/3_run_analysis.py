"""
Paso 3 — Análisis: gráficas, métricas y reporte HTML
=====================================================
Lee la tabla combinada (producida por el paso 2) y genera:
  - Figuras scatter + Bland-Altman por grupo (normal + versión con etiquetas)
  - Heatmaps resumen de concordancia (SUV y DIXON)
  - Excel con métricas (Metricas_Manual_vs_Auto.xlsx)
  - Reporte HTML interactivo (Report_Manual_vs_Auto.html)

Entrada : RESULTS/Combined_Measurements.xlsx
Salidas : RESULTS/Fig_*.png  (normal + _labeled)
          RESULTS/Fig_Summary_SUV.png / Fig_Summary_DIXON.png
          RESULTS/Metricas_Manual_vs_Auto.xlsx
          RESULTS/Report_Manual_vs_Auto.html
"""

import sys
import os
import base64
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import colorsys
from scipy import stats
import pingouin as pg
from datetime import date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from nvitk.core.logger import Logger

warnings.filterwarnings("ignore")

sys.path.insert(0, __file__.rsplit("\\", 1)[0])
from config import COMBINED_TABLE_PATH, METRICS_PATH, REPORT_PATH, output_dir

# This script writes figures and reports, so it is the thing that should create the output
# directory. Importing the config module no longer does it as a side effect.
OUT_DIR = output_dir(create=True)

log = Logger()

# ─────────────────────────────────────────────────────────────────────────────
# Estilo gráfico
# ─────────────────────────────────────────────────────────────────────────────

PALETTE = {
    "scatter": "#2E86AB", "identity": "#E84855", "regression": "#F4A261",
    "ba_mean": "#2E86AB", "ba_loa":   "#E84855", "ba_zero":    "#555555",
    "grid":    "#E8E8E8",
}
plt.rcParams.update({
    "font.family": "DejaVu Sans", "font.size": 10,
    "axes.titlesize": 11, "axes.titleweight": "bold", "axes.labelsize": 10,
    "axes.spines.top": False, "axes.spines.right": False,
    "axes.grid": True, "axes.grid.which": "major",
    "grid.color": PALETTE["grid"], "grid.linewidth": 0.8,
    "figure.dpi": 150, "savefig.dpi": 200, "savefig.bbox": "tight",
})

def darken_hex(hex_color, factor=0.8):
    """Darken a ``#rrggbb`` hex color by scaling its HLS lightness by *factor*."""
    hex_color = hex_color.lstrip('#')
    
    # HEX → RGB (0-255)
    r, g, b = [int(hex_color[i:i+2], 16) for i in (0, 2, 4)]
    
    # RGB → HLS (0-1)
    r, g, b = r/255, g/255, b/255
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    
    # Reducir luminosidad
    l *= factor
    
    # Volver a RGB
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    r, g, b = int(r*255), int(g*255), int(b*255)
    
    # RGB → HEX
    return f"#{r:02x}{g:02x}{b:02x}"


# ─────────────────────────────────────────────────────────────────────────────
# Pares de comparación por grupo/figura
# ─────────────────────────────────────────────────────────────────────────────

COMPARISONS = {

    "SUV_Medula": {
        "title": "SUV \u2013 M\u00e9dula \u00d3sea Vertebral L3/L4  (n\u224822)",
        "fig_name": "Fig_SUV_Medula",
        "modality": "SUV",
        "pairs": [
            {"label": "L3 SUVmax",  "manual": "man_L3_SUVmax",  "auto": "L3_SUVMAX",  "units": "SUV"},
            {"label": "L3 SUVmean", "manual": "man_L3_SUVmean", "auto": "L3_SUVmean", "units": "SUV"},
            {"label": "L4 SUVmax",  "manual": "man_L4_SUVmax",  "auto": "L4_SUVMAX",  "units": "SUV"},
            {"label": "L4 SUVmean", "manual": "man_L4_SUVmean", "auto": "L4_SUVmean", "units": "SUV"},
            {"label": "M\u00e9dula L3-L4\n(media SUVmax)",
             "manual": "man_Medula_SUVmax_avg", "auto": "auto_Medula_SUVmax_avg",
             "units": "SUV", "note": True,
             "xlabel": "Manual", "ylabel": "Autom\u00e1tico [NOTA 1]"},
        ],
    },

    "SUV_Higado": {
        "title": "SUV \u2013 H\u00edgado  (n=6, semana 1)",
        "fig_name": "Fig_SUV_Higado",
        "modality": "SUV",
        "pairs": [
            {"label": "H\u00edgado SUVmax",  "manual": "man2_Higado_SUVmax", "auto": "HIGADO_SUVMAX",  "units": "SUV", "xlabel": "Manual (n=6)"},
            {"label": "H\u00edgado SUVmean", "manual": "man2_Higado_SUVmed", "auto": "HIGADO_SUVmean", "units": "SUV", "xlabel": "Manual (n=6)"},
        ],
    },

    "SUV_Bazo": {
        "title": "SUV \u2013 Bazo  (n=6, semana 1)",
        "fig_name": "Fig_SUV_Bazo",
        "modality": "SUV",
        "pairs": [
            {"label": "Bazo SUVmax",  "manual": "man2_Bazo_SUVmax", "auto": "BAZO_SUVMAX",  "units": "SUV", "xlabel": "Manual (n=6)"},
            {"label": "Bazo SUVmean", "manual": "man2_Bazo_SUVmed", "auto": "BAZO_SUVmean", "units": "SUV", "xlabel": "Manual (n=6)"},
        ],
    },

    "SUV_Pancreas": {
        "title": "SUV \u2013 P\u00e1ncreas  (n=6, semana 1)",
        "fig_name": "Fig_SUV_Pancreas",
        "modality": "SUV",
        "pairs": [
            {"label": "P\u00e1ncreas SUVmax",  "manual": "man2_Pancreas_SUVmax", "auto": "PANCREAS_SUVMAX",  "units": "SUV", "xlabel": "Manual (n=6)"},
            {"label": "P\u00e1ncreas SUVmean", "manual": "man2_Pancreas_SUVmed", "auto": "PANCREAS_SUVmean", "units": "SUV", "xlabel": "Manual (n=6)"},
        ],
    },

    "SUV_Cuadriceps": {
        "title": "SUV \u2013 Cu\u00e1driceps  (n=6, semana 1)",
        "fig_name": "Fig_SUV_Cuadriceps",
        "modality": "SUV",
        "pairs": [
            {"label": "Cu\u00e1d. DER\nSUVmax",  "manual": "man2_Cuad_DER_SUVmax", "auto": "CUADRICEPS_R_SUVMAX",  "units": "SUV", "xlabel": "Manual DER (n=6)"},
            {"label": "Cu\u00e1d. DER\nSUVmean", "manual": "man2_Cuad_DER_SUVmed", "auto": "CUADRICEPS_R_SUVmean", "units": "SUV", "xlabel": "Manual DER (n=6)"},
            {"label": "Cu\u00e1d. IZQ\nSUVmax",  "manual": "man2_Cuad_IZQ_SUVmax", "auto": "CUADRICEPS_L_SUVMAX",  "units": "SUV", "xlabel": "Manual IZQ (n=6)"},
            {"label": "Cu\u00e1d. IZQ\nSUVmean", "manual": "man2_Cuad_IZQ_SUVmed", "auto": "CUADRICEPS_L_SUVmean", "units": "SUV", "xlabel": "Manual IZQ (n=6)"},
        ],
    },

    "SUV_Paravertebral": {
        "title": "SUV \u2013 Musculatura Paravertebral  (n=6, semana 1)",
        "fig_name": "Fig_SUV_Paravertebral",
        "modality": "SUV",
        "pairs": [
            {"label": "Paravert. DER\nSUVmax",  "manual": "man2_Paravert_DER_SUVmax", "auto": "PARAVERTEBRAL_R_SUVMAX",  "units": "SUV", "xlabel": "Manual DER (n=6)"},
            {"label": "Paravert. DER\nSUVmean", "manual": "man2_Paravert_DER_SUVmed", "auto": "PARAVERTEBRAL_R_SUVmean", "units": "SUV", "xlabel": "Manual DER (n=6)"},
            {"label": "Paravert. IZQ\nSUVmax",  "manual": "man2_Paravert_IZQ_SUVmax", "auto": "PARAVERTEBRAL_L_SUVMAX",  "units": "SUV", "xlabel": "Manual IZQ (n=6)"},
            {"label": "Paravert. IZQ\nSUVmean", "manual": "man2_Paravert_IZQ_SUVmed", "auto": "PARAVERTEBRAL_L_SUVmean", "units": "SUV", "xlabel": "Manual IZQ (n=6)"},
        ],
    },

    "DIXON_Higado": {
        "title": "DIXON \u2013 H\u00edgado  (n\u224822)",
        "fig_name": "Fig_DIXON_Higado",
        "modality": "DIXON",
        "pairs": [
            {"label": "Liver Fat Fraction", "manual": "man_Liver_FF",  "auto": "DIXON_LIVER_FF",  "units": "%"},
            {"label": "Liver Volume",        "manual": "man_Liver_VOL", "auto": "DIXON_LIVER_VOL", "units": "cc"},
            {"label": "Liver T2*",           "manual": "man_Liver_T2",  "auto": "DIXON_LIVER_T2",  "units": "ms"},
            {"label": "Liver R2 \u26a0",     "manual": "man_Liver_R2",  "auto": "DIXON_LIVER_R2",
             "units": "\u00bfms\u207b\u00b9?", "note": True,
             "xlabel": "Manual (escala desconocida)", "ylabel": "Autom\u00e1tico (\u00bfms\u207b\u00b9?)"},
        ],
    },

    "DIXON_Rinon": {
        "title": "DIXON \u2013 Ri\u00f1\u00f3n  (n=6, semana 1)",
        "fig_name": "Fig_DIXON_Rinon",
        "modality": "DIXON",
        "pairs": [
            {"label": "Ri\u00f1\u00f3n DER\nFF",  "manual": "man2_Rinon_DER_FF", "auto": "DIXON_KIDNEY_R_FF", "units": "%",  "xlabel": "Manual DER (n=6)"},
            {"label": "Ri\u00f1\u00f3n DER\nT2*", "manual": "man2_Rinon_DER_T2", "auto": "DIXON_KIDNEY_R_T2", "units": "ms", "xlabel": "Manual DER (n=6)"},
            {"label": "Ri\u00f1\u00f3n IZQ\nFF",  "manual": "man2_Rinon_IZQ_FF", "auto": "DIXON_KIDNEY_L_FF", "units": "%",  "xlabel": "Manual IZQ (n=6)"},
            {"label": "Ri\u00f1\u00f3n IZQ\nT2*", "manual": "man2_Rinon_IZQ_T2", "auto": "DIXON_KIDNEY_L_T2", "units": "ms", "xlabel": "Manual IZQ (n=6)"},
        ],
    },

    "DIXON_Pancreas": {
        "title": "DIXON \u2013 P\u00e1ncreas  (n=6, semana 1)",
        "fig_name": "Fig_DIXON_Pancreas",
        "modality": "DIXON",
        "pairs": [
            {"label": "P\u00e1ncreas FF",  "manual": "man2_Pancreas_FF", "auto": "DIXON_PANCREAS_FF", "units": "%",  "xlabel": "Manual (n=6)"},
            {"label": "P\u00e1ncreas T2*", "manual": "man2_Pancreas_T2", "auto": "DIXON_PANCREAS_T2", "units": "ms", "xlabel": "Manual (n=6)"},
        ],
    },

    "DIXON_Musculo": {
        "title": "DIXON \u2013 Cu\u00e1driceps  (n=6, semana 1)",
        "fig_name": "Fig_DIXON_Musculo",
        "modality": "DIXON",
        "pairs": [
            {"label": "Cu\u00e1d. DER\nFF",  "manual": "man2_Cuad_DER_FF", "auto": "DIXON_L_QM_R_FF", "units": "%",  "xlabel": "Manual DER (n=6)"},
            {"label": "Cu\u00e1d. DER\nT2*", "manual": "man2_Cuad_DER_T2", "auto": "DIXON_L_QM_R_T2", "units": "ms", "xlabel": "Manual DER (n=6)"},
            {"label": "Cu\u00e1d. IZQ\nFF",  "manual": "man2_Cuad_IZQ_FF", "auto": "DIXON_L_QM_L_FF", "units": "%",  "xlabel": "Manual IZQ (n=6)"},
            {"label": "Cu\u00e1d. IZQ\nT2*", "manual": "man2_Cuad_IZQ_T2", "auto": "DIXON_L_QM_L_T2", "units": "ms", "xlabel": "Manual IZQ (n=6)"},
        ],
    },

    "DIXON_Medula": {
        "title": "DIXON \u2013 M\u00e9dula \u00d3sea L3/L4  (n=6, semana 1)",
        "fig_name": "Fig_DIXON_Medula",
        "modality": "DIXON",
        "pairs": [
            {"label": "M\u00e9dula L3 FF",  "manual": "man2_BN_L3_FF", "auto": "DIXON_BN_L3_FF", "units": "%",  "xlabel": "Manual (n=6)"},
            {"label": "M\u00e9dula L3 T2*", "manual": "man2_BN_L3_T2", "auto": "DIXON_BN_L3_T2", "units": "ms", "xlabel": "Manual (n=6)"},
            {"label": "M\u00e9dula L4 FF",  "manual": "man2_BN_L4_FF", "auto": "DIXON_BN_L4_FF", "units": "%",  "xlabel": "Manual (n=6)"},
            {"label": "M\u00e9dula L4 T2*", "manual": "man2_BN_L4_T2", "auto": "DIXON_BN_L4_T2", "units": "ms", "xlabel": "Manual (n=6)"},
        ],
    },

    "DIXON_Volumenes": {
        "title": "DIXON \u2013 Volumen manual (cc) vs VOL autom\u00e1tico (cc)",
        "fig_name": "Fig_DIXON_Volumenes",
        "modality": "DIXON",
        "area_figure": True,   # filtrar pares con n=0
        "pairs": [
            {"label": "Ri\u00f1\u00f3n DER\nVOL (cc)",       "manual": "man2_Rinon_DER_VOL", "auto": "DIXON_KIDNEY_R_VOL",  "units": "cc", "xlabel": "Manual DER (cc)",       "ylabel": "Auto (cc)"},
            {"label": "Ri\u00f1\u00f3n IZQ\nVOL (cc)",       "manual": "man2_Rinon_IZQ_VOL", "auto": "DIXON_KIDNEY_L_VOL",  "units": "cc", "xlabel": "Manual IZQ (cc)",       "ylabel": "Auto (cc)"},
            {"label": "Cu\u00e1d. DER\nVOL (cc)",            "manual": "man2_Cuad_DER_VOL",  "auto": "DIXON_L_QM_R_VOL",    "units": "cc", "xlabel": "Manual DER (cc)",       "ylabel": "Auto (cc)"},
            {"label": "Cu\u00e1d. IZQ\nVOL (cc)",            "manual": "man2_Cuad_IZQ_VOL",  "auto": "DIXON_L_QM_L_VOL",    "units": "cc", "xlabel": "Manual IZQ (cc)",       "ylabel": "Auto (cc)"},
            {"label": "P\u00e1ncreas\nVOL (cc)",             "manual": "man2_Pancreas_VOL",  "auto": "DIXON_PANCREAS_VOL",  "units": "cc", "xlabel": "Manual (cc)",           "ylabel": "Auto (cc)"},
            {"label": "M\u00e9dula L3\nVOL (cc)",            "manual": "man2_BN_L3_VOL",     "auto": "DIXON_BN_L3_VOL",     "units": "cc", "xlabel": "Manual (cc)",           "ylabel": "Auto (cc)"},
            {"label": "M\u00e9dula L4\nVOL (cc)",            "manual": "man2_BN_L4_VOL",     "auto": "DIXON_BN_L4_VOL",     "units": "cc", "xlabel": "Manual (cc)",           "ylabel": "Auto (cc)"},
        ],
    },
}

# Pares para la hoja "Datos_Comparativos" del Excel
PAIR_COLS = [
    ("man_L3_SUVmax",            "L3_SUVMAX",                 "L3 SUVmax",             "SUV"),
    ("man_L3_SUVmean",           "L3_SUVmean",                "L3 SUVmean",            "SUV"),
    ("man_L4_SUVmax",            "L4_SUVMAX",                 "L4 SUVmax",             "SUV"),
    ("man_L4_SUVmean",           "L4_SUVmean",                "L4 SUVmean",            "SUV"),
    ("man_Medula_SUVmax_avg",    "auto_Medula_SUVmax_avg",    "Medula SUVmax",         "SUV"),
    ("man2_Higado_SUVmax",       "HIGADO_SUVMAX",             "Higado SUVmax",         "SUV"),
    ("man2_Higado_SUVmed",       "HIGADO_SUVmean",            "Higado SUVmean",        "SUV"),
    ("man2_Bazo_SUVmax",         "BAZO_SUVMAX",               "Bazo SUVmax",           "SUV"),
    ("man2_Bazo_SUVmed",         "BAZO_SUVmean",              "Bazo SUVmean",          "SUV"),
    ("man2_Pancreas_SUVmax",     "PANCREAS_SUVMAX",           "Pancreas SUVmax",       "SUV"),
    ("man2_Pancreas_SUVmed",     "PANCREAS_SUVmean",          "Pancreas SUVmean",      "SUV"),
    ("man2_Cuad_DER_SUVmax",     "CUADRICEPS_R_SUVMAX",       "Cuad DER SUVmax",       "SUV"),
    ("man2_Cuad_DER_SUVmed",     "CUADRICEPS_R_SUVmean",      "Cuad DER SUVmean",      "SUV"),
    ("man2_Cuad_IZQ_SUVmax",     "CUADRICEPS_L_SUVMAX",       "Cuad IZQ SUVmax",       "SUV"),
    ("man2_Cuad_IZQ_SUVmed",     "CUADRICEPS_L_SUVmean",      "Cuad IZQ SUVmean",      "SUV"),
    ("man2_Paravert_DER_SUVmax", "PARAVERTEBRAL_R_SUVMAX",    "Paravert DER SUVmax",   "SUV"),
    ("man2_Paravert_DER_SUVmed", "PARAVERTEBRAL_R_SUVmean",   "Paravert DER SUVmean",  "SUV"),
    ("man2_Paravert_IZQ_SUVmax", "PARAVERTEBRAL_L_SUVMAX",    "Paravert IZQ SUVmax",   "SUV"),
    ("man2_Paravert_IZQ_SUVmed", "PARAVERTEBRAL_L_SUVmean",   "Paravert IZQ SUVmean",  "SUV"),
    ("man_Liver_FF",             "DIXON_LIVER_FF",            "Liver FF",              "%"),
    ("man_Liver_VOL",            "DIXON_LIVER_VOL",           "Liver VOL",             "cc"),
    ("man_Liver_T2",             "DIXON_LIVER_T2",            "Liver T2*",             "ms"),
    ("man_Liver_R2",             "DIXON_LIVER_R2",            "Liver R2",              "?"),
    ("man2_Rinon_DER_FF",        "DIXON_KIDNEY_R_FF",         "Rinon DER FF",          "%"),
    ("man2_Rinon_DER_T2",        "DIXON_KIDNEY_R_T2",         "Rinon DER T2*",         "ms"),
    ("man2_Rinon_IZQ_FF",        "DIXON_KIDNEY_L_FF",         "Rinon IZQ FF",          "%"),
    ("man2_Rinon_IZQ_T2",        "DIXON_KIDNEY_L_T2",         "Rinon IZQ T2*",         "ms"),
    ("man2_Pancreas_FF",         "DIXON_PANCREAS_FF",         "Pancreas FF",           "%"),
    ("man2_Pancreas_T2",         "DIXON_PANCREAS_T2",         "Pancreas T2*",          "ms"),
    ("man2_Cuad_DER_FF",         "DIXON_L_QM_R_FF",           "Cuad DER FF",           "%"),
    ("man2_Cuad_DER_T2",         "DIXON_L_QM_R_T2",           "Cuad DER T2*",          "ms"),
    ("man2_Cuad_IZQ_FF",         "DIXON_L_QM_L_FF",           "Cuad IZQ FF",           "%"),
    ("man2_Cuad_IZQ_T2",         "DIXON_L_QM_L_T2",           "Cuad IZQ T2*",          "ms"),
    ("man2_BN_L3_FF",            "DIXON_BN_L3_FF",            "Medula L3 FF",          "%"),
    ("man2_BN_L3_T2",            "DIXON_BN_L3_T2",            "Medula L3 T2*",         "ms"),
    ("man2_BN_L4_FF",            "DIXON_BN_L4_FF",            "Medula L4 FF",          "%"),
    ("man2_BN_L4_T2",            "DIXON_BN_L4_T2",            "Medula L4 T2*",         "ms"),
    ("man2_Rinon_DER_VOL",       "DIXON_KIDNEY_R_VOL",        "Rinon DER VOL",         "cc"),
    ("man2_Rinon_IZQ_VOL",       "DIXON_KIDNEY_L_VOL",        "Rinon IZQ VOL",         "cc"),
    ("man2_Cuad_DER_VOL",        "DIXON_L_QM_R_VOL",          "Cuad DER VOL",          "cc"),
    ("man2_Cuad_IZQ_VOL",        "DIXON_L_QM_L_VOL",          "Cuad IZQ VOL",          "cc"),
    ("man2_Pancreas_VOL",        "DIXON_PANCREAS_VOL",        "Pancreas VOL",          "cc"),
    ("man2_BN_L3_VOL",           "DIXON_BN_L3_VOL",           "Medula L3 VOL",         "cc"),
    ("man2_BN_L4_VOL",           "DIXON_BN_L4_VOL",           "Medula L4 VOL",         "cc"),
]


# ─────────────────────────────────────────────────────────────────────────────
# Métricas de concordancia
# ─────────────────────────────────────────────────────────────────────────────

def compute_ccc(x, y):
    """Lin's concordance correlation coefficient for *x* vs *y*, with a 95% Fisher-z CI."""
    n = len(x)
    mx, my = np.mean(x), np.mean(y)
    sx2 = np.var(x, ddof=1); sy2 = np.var(y, ddof=1)
    sxy = np.cov(x, y, ddof=1)[0, 1]
    ccc = (2 * sxy) / (sx2 + sy2 + (mx - my) ** 2)
    zc  = np.arctanh(np.clip(ccc, -0.9999, 0.9999))
    se  = 1.0 / np.sqrt(n - 3)
    return float(ccc), float(np.tanh(zc - 1.96*se)), float(np.tanh(zc + 1.96*se))


_NAN_DICT = {k: np.nan for k in [
    "n", "r", "r_p", "rho", "rho_p",
    "ccc", "ccc_lo", "ccc_hi",
    "icc", "icc_lo", "icc_hi",
    "bias", "loa_lo", "loa_hi", "rmse", "mae",
]}


def compute_metrics(manual_vals, auto_vals):
    """Concordance metrics (Pearson r, Spearman rho, CCC, ICC(A,1), Bland-Altman bias/LoA,
    RMSE, MAE) between manual and automatic measurements; all-NaN dict if n < 4."""
    df_pair = pd.DataFrame({"m": manual_vals, "a": auto_vals}).dropna()
    n = len(df_pair)
    out = {**_NAN_DICT, "n": n}
    if n < 4:
        return out
    x, y = df_pair["m"].values, df_pair["a"].values
    out["r"],   out["r_p"]   = stats.pearsonr(x, y)
    out["rho"], out["rho_p"] = stats.spearmanr(x, y)
    out["ccc"], out["ccc_lo"], out["ccc_hi"] = compute_ccc(x, y)
    icc_df = pg.intraclass_corr(
        data=pd.DataFrame({"targets": np.arange(n).tolist() * 2,
                           "raters":  ["Manual"] * n + ["Auto"] * n,
                           "ratings": list(x) + list(y)}),
        targets="targets", raters="raters", ratings="ratings",
    )
    row = icc_df[icc_df["Type"] == "ICC(A,1)"].iloc[0]
    out["icc"], out["icc_lo"], out["icc_hi"] = row["ICC"], row["CI95"][0], row["CI95"][1]
    diff = x - y
    out["bias"]   = float(np.mean(diff))
    sd            = float(np.std(diff, ddof=1))
    out["loa_lo"] = out["bias"] - 1.96 * sd
    out["loa_hi"] = out["bias"] + 1.96 * sd
    out["rmse"]   = float(np.sqrt(np.mean(diff ** 2)))
    out["mae"]    = float(np.mean(np.abs(diff)))
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Funciones de graficado
# ─────────────────────────────────────────────────────────────────────────────

def _fmt(v):
    """Format *v* to 3 decimals, or ``"n/d"`` if None/NaN."""
    return f"{float(v):.3f}" if (v is not None and not np.isnan(float(v))) else "n/d"


def _best_legend_loc(xv, yv, lim):
    """Legend corner (of the 4 quadrants) with the fewest scatter points, to avoid overlap."""
    mid = (lim[0] + lim[1]) / 2
    counts = {
        "upper left":  np.sum((xv < mid) & (yv >= mid)),
        "upper right": np.sum((xv >= mid) & (yv >= mid)),
        "lower left":  np.sum((xv < mid) & (yv < mid)),
        "lower right": np.sum((xv >= mid) & (yv < mid)),
    }
    return min(counts, key=counts.get)


def scatter_panel(ax, x, y, label, units, metrics, note=False,
                  xlabel=None, ylabel=None, point_labels=None):
    """Draw a manual-vs-automatic scatter panel on *ax* with identity/regression lines and a
    concordance-metrics text box (handles n=0/1 edge cases separately)."""
    from matplotlib.lines import Line2D
    mask = ~(np.isnan(x) | np.isnan(y))
    xv, yv = x[mask], y[mask]
    lv = np.array(point_labels)[mask] if point_labels is not None else None
    n_valid = len(xv)

    if n_valid < 1:
        ax.text(0.5, 0.5, "Sin datos\n(n=0)", transform=ax.transAxes,
                ha="center", va="center", fontsize=11, color="#C0392B")
        ax.set_title(label + (" \u26a0" if note else ""), fontsize=10)
        return

    ax.scatter(xv, yv, color=PALETTE["scatter"], s=60, alpha=0.8,
               edgecolors="white", linewidths=0.5, zorder=3)
    if lv is not None:
        for xi, yi, li in zip(xv, yv, lv):
            ax.annotate(str(li), (xi, yi), textcoords="offset points", xytext=(4, 4),
                        fontsize=6.5, color="#333333", zorder=4)

    if n_valid == 1:
        cx, cy = float(xv[0]), float(yv[0])
        half = max(abs(cx) * 0.5, abs(cy) * 0.5, 1.0)
        lim  = (min(cx, cy) - half, max(cx, cy) + half)
        ax.set_xlim(lim); ax.set_ylim(lim); ax.set_aspect("equal")
        ax.plot(lim, lim, color=PALETTE["identity"], lw=1.5, ls="--", zorder=2)
        ax.text(0.04, 0.97, f"n={n_valid}\n(n insuficiente para m\u00e9tricas)",
                transform=ax.transAxes, va="top", ha="left", fontsize=8, linespacing=1.6,
                bbox=dict(boxstyle="round,pad=0.35", fc="white", alpha=0.88, ec="none"))
        ax.set_title(label + (" \u26a0" if note else ""), fontsize=10)
        ax.set_xlabel(xlabel if xlabel else f"Manual ({units})")
        ax.set_ylabel(ylabel if ylabel else f"Autom\u00e1tico ({units})")
        return

    lim = (min(xv.min(), yv.min()), max(xv.max(), yv.max()))
    pad = (lim[1] - lim[0]) * 0.05
    lim = (lim[0] - pad, lim[1] + pad)
    ax.plot(lim, lim, color=PALETTE["identity"], lw=1.5, ls="--", zorder=2)

    m_reg = b_reg = np.nan
    if n_valid >= 3:
        m_reg, b_reg, *_ = stats.linregress(xv, yv)
        xs = np.linspace(lim[0], lim[1], 200)
        ax.plot(xs, m_reg * xs + b_reg, color=PALETTE["regression"], lw=2, zorder=2)

    ax.set_xlim(lim); ax.set_ylim(lim); ax.set_aspect("equal")

    has_m = not np.isnan(metrics.get("r", np.nan))
    if has_m:
        p_r   = metrics["r_p"]
        p_str = "<.001" if p_r < 0.001 else f"={p_r:.3f}"
        reg_str = f"y={m_reg:.2f}x{b_reg:+.2f}" if not np.isnan(m_reg) else ""
        lines = [
            f"n={int(metrics['n'])}",
            f"r={metrics['r']:.3f} (p{p_str})",
            f"rho={metrics['rho']:.3f}",
            f"CCC={_fmt(metrics['ccc'])} [{_fmt(metrics['ccc_lo'])},{_fmt(metrics['ccc_hi'])}]",
            f"ICC={_fmt(metrics['icc'])} [{_fmt(metrics['icc_lo'])},{_fmt(metrics['icc_hi'])}]",
            reg_str,
        ]
    else:
        lines = [f"n={n_valid}", "(n insuficiente para m\u00e9tricas)"]
    ax.text(0.04, 0.97, "\n".join(l for l in lines if l),
            transform=ax.transAxes, va="top", ha="left", fontsize=8, linespacing=1.6,
            bbox=dict(boxstyle="round,pad=0.35", fc="white", alpha=0.88, ec="none"))

    leg_loc = _best_legend_loc(xv, yv, lim)
    ax.legend(handles=[
        Line2D([0], [0], color=PALETTE["identity"],   lw=1.5, ls="--", label="Identidad"),
        Line2D([0], [0], color=PALETTE["regression"], lw=2,   ls="-",  label="Regresi\u00f3n"),
    ], fontsize=7, framealpha=0.85, loc=leg_loc, handlelength=1.5, borderaxespad=0.5)

    ax.set_title(label + (" \u26a0" if note else ""), fontsize=10)
    ax.set_xlabel(xlabel if xlabel else f"Manual ({units})")
    ax.set_ylabel(ylabel if ylabel else f"Autom\u00e1tico ({units})")


def bland_altman_panel(ax, x, y, label, units, metrics, note=False, point_labels=None):
    """Draw a Bland-Altman panel on *ax* (mean vs. difference) with bias/LoA lines annotated
    (handles n=0/1 edge cases separately)."""
    mask = ~(np.isnan(x) | np.isnan(y))
    xv, yv = x[mask], y[mask]
    lv = np.array(point_labels)[mask] if point_labels is not None else None
    n_valid = len(xv)

    if n_valid < 1:
        ax.text(0.5, 0.5, "Sin datos\n(n=0)", transform=ax.transAxes,
                ha="center", va="center", fontsize=11, color="#C0392B")
        ax.set_title(label + (" \u26a0" if note else ""), fontsize=10)
        return

    means = (xv + yv) / 2
    diffs = xv - yv

    if n_valid == 1:
        ax.scatter(means, diffs, color=PALETTE["scatter"], s=60, alpha=0.8,
                   edgecolors="white", linewidths=0.5, zorder=3)
        if lv is not None:
            ax.annotate(str(lv[0]), (float(means[0]), float(diffs[0])),
                        textcoords="offset points", xytext=(4, 4), fontsize=6.5, zorder=4)
        ax.axhline(0, color=PALETTE["ba_zero"], lw=1, ls=":", zorder=2)
        half = max(abs(float(means[0])) * 0.5, 1.0)
        ax.set_xlim(float(means[0]) - half, float(means[0]) + half)
        ax.text(0.04, 0.97, f"n={n_valid}\n(n insuficiente para m\u00e9tricas)",
                transform=ax.transAxes, va="top", ha="left", fontsize=8, linespacing=1.6,
                bbox=dict(boxstyle="round,pad=0.35", fc="white", alpha=0.88, ec="none"))
        ax.set_title(label + (" \u26a0" if note else ""), fontsize=10)
        ax.set_xlabel(f"Media (Manual + Autom\u00e1tico) / 2  ({units})")
        ax.set_ylabel(f"Diferencia  Manual \u2212 Autom\u00e1tico  ({units})")
        return

    ax.scatter(means, diffs, color=PALETTE["scatter"], s=50, alpha=0.8,
               edgecolors="white", linewidths=0.5, zorder=3)
    if lv is not None:
        for mi, di, li in zip(means, diffs, lv):
            ax.annotate(str(li), (mi, di), textcoords="offset points", xytext=(4, 4),
                        fontsize=6.5, color="#333333", zorder=4)

    has_m = not np.isnan(metrics.get("bias", np.nan))
    pad_x = (means.max() - means.min()) * 0.05 if means.max() > means.min() else 1.0
    xline = np.array([means.min() - pad_x, means.max() + pad_x])

    if has_m:
        bias, loa_lo, loa_hi = metrics["bias"], metrics["loa_lo"], metrics["loa_hi"]
        ax.axhline(bias,   color=PALETTE["ba_mean"], lw=2,   zorder=2)
        ax.axhline(loa_lo, color=PALETTE["ba_loa"],  lw=1.5, ls="--", zorder=2)
        ax.axhline(loa_hi, color=PALETTE["ba_loa"],  lw=1.5, ls="--", zorder=2)
        ax.fill_between(xline, loa_lo, loa_hi, alpha=0.08, color=PALETTE["ba_mean"], zorder=1)
        x_ann = means.max() + pad_x * 0.3
        ax.annotate(f"Bias={bias:.3f}",         xy=(x_ann, bias),   fontsize=7.5, color=darken_hex(PALETTE["ba_mean"],0.7), va="center", ha="left", fontweight="bold")
        ax.annotate(f"LoA+={loa_hi:.3f}",       xy=(x_ann, loa_hi), fontsize=7.5, color=darken_hex(PALETTE["ba_loa"],0.7),  va="center", ha="left")
        ax.annotate(f"LoA\u2212={loa_lo:.3f}",  xy=(x_ann, loa_lo), fontsize=7.5, color=darken_hex(PALETTE["ba_loa"],0.7),  va="center", ha="left")
    else:
        ax.axhline(np.mean(diffs), color=PALETTE["ba_mean"], lw=1.5, ls=":", zorder=2)

    ax.axhline(0, color=PALETTE["ba_zero"], lw=1, ls=":", zorder=2)
    ax.set_title(label + (" \u26a0" if note else ""), fontsize=10)
    ax.set_xlabel(f"Media (Manual + Autom\u00e1tico) / 2  ({units})")
    ax.set_ylabel(f"Diferencia  Manual \u2212 Autom\u00e1tico  ({units})")
    ax.set_xlim(xline[0], xline[1] + (pad_x * 3.5 if has_m else 0))


# ─────────────────────────────────────────────────────────────────────────────
# Heatmap resumen
# ─────────────────────────────────────────────────────────────────────────────

def make_heatmap(summary_df, title, out_path):
    """Save a combined correlation/concordance heatmap + Bland-Altman summary table figure."""
    metric_cols = ["r (Pearson)", "rho (Spearman)", "CCC", "ICC (2,1)"]
    heat_labels = ["Pearson r", "Spearman \u03c1", "CCC", "ICC(2,1)"]
    heat_data   = summary_df.set_index("Variable")[metric_cols].astype(float)

    fig, (ax_heat, ax_tbl) = plt.subplots(
        1, 2, figsize=(16, max(6, len(heat_data) * 0.55 + 2)),
        gridspec_kw={"width_ratios": [1.6, 1], "wspace": 0.5},
    )
    norm = mcolors.Normalize(vmin=0, vmax=1)
    im   = ax_heat.imshow(heat_data.values, cmap=plt.cm.RdYlGn, norm=norm, aspect="auto")
    plt.colorbar(im, ax=ax_heat, fraction=0.025, pad=0.02, label="Valor (0\u20131)")

    ax_heat.set_xticks(range(len(heat_labels))); ax_heat.set_xticklabels(heat_labels, fontsize=9.5)
    ax_heat.set_yticks(range(len(heat_data)));   ax_heat.set_yticklabels(heat_data.index, fontsize=8)
    for ri in range(heat_data.shape[0]):
        for ci in range(heat_data.shape[1]):
            val = heat_data.values[ri, ci]
            if np.isnan(val):
                ax_heat.text(ci, ri, "n/d", ha="center", va="center", fontsize=8)
            else:
                tc = "white" if (val < 0.35 or val > 0.82) else "black"
                ax_heat.text(ci, ri, f"{val:.3f}", ha="center", va="center",
                             fontsize=8.5, fontweight="bold", color=tc)
    ax_heat.set_title("Correlaci\u00f3n y concordancia", fontsize=11, fontweight="bold")
    ax_heat.spines[:].set_visible(False); ax_heat.grid(False)

    ba_cols  = ["Variable", "n", "Bias", "LoA inf", "LoA sup", "RMSE"]
    tbl_data = summary_df[ba_cols].copy()
    for c in ["Bias", "LoA inf", "LoA sup", "RMSE"]:
        tbl_data[c] = tbl_data[c].apply(
            lambda v: f"{v:.3f}" if isinstance(v, float) and not np.isnan(v) else "\u2013"
        )
    ax_tbl.axis("off")
    tbl = ax_tbl.table(cellText=tbl_data.values, colLabels=ba_cols, cellLoc="center", loc="center")
    tbl.auto_set_font_size(False); tbl.set_fontsize(8)
    tbl.auto_set_column_width(col=list(range(len(ba_cols))))
    for j in range(len(ba_cols)):
        tbl[0, j].set_facecolor("#2E4057"); tbl[0, j].set_text_props(color="white", fontweight="bold")
    for i in range(1, len(tbl_data) + 1):
        c = "#F7F9FC" if i % 2 == 0 else "white"
        for j in range(len(ba_cols)):
            tbl[i, j].set_facecolor(c)
    ax_tbl.set_title("Bland-Altman", fontsize=11, fontweight="bold", pad=10)

    fig.suptitle(title, fontsize=13, fontweight="bold", y=1.02)
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)
    log.info(f"  OK {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# Excel de métricas
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(summary, comp_table, out_path):
    """Write the formatted metrics + comparative-data Excel workbook (colored headers,
    zebra rows, conditional formatting on Diff% columns)."""
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    _GRP_COLORS = {
        "L3": "AED6F1", "L4": "A9DFBF", "Medula SUVmax": "D7BDE2",
        "Higado SUV": "FADBD8", "Bazo SUV": "FADBD8", "Pancreas SUV": "FADBD8",
        "Cuad BIL SUV": "FDEBD0", "Paravert BIL SUV": "FDEBD0",
        "Liver": "FAD7A0",
        "Rinon": "D5F5E3", "Pancreas FF": "D6EAF8", "Pancreas T2": "D6EAF8",
        "Cuad BIL F": "E8DAEF", "Cuad BIL T": "E8DAEF",
        "Medula L3": "EBF5FB", "Medula L4": "F0F3F4",
    }

    def _header_color(col_name):
        """Fill color for a comparative-table header whose name contains a known group keyword."""
        for kw, color in _GRP_COLORS.items():
            if kw in col_name:
                return color
        return "D5D8DC"

    num_cols = ["r (Pearson)", "rho (Spearman)", "CCC", "ICC (2,1)",
                "Bias", "LoA inf", "LoA sup", "RMSE", "MAE"]
    export_df = summary.copy()
    for col in num_cols:
        export_df[col] = export_df[col].apply(
            lambda v: round(float(v), 4)
            if isinstance(v, (float, np.floating)) and not np.isnan(v) else v
        )

    hdr_fill  = PatternFill("solid", fgColor="2E4057")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    alt_fill1 = PatternFill("solid", fgColor="F7F9FC")
    alt_fill2 = PatternFill("solid", fgColor="FFFFFF")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Hoja 1: Métricas
        export_df.to_excel(writer, index=False, sheet_name="Metricas")
        ws = writer.sheets["Metricas"]
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 28
        prev_grp = None
        for i, row_cells in enumerate(ws.iter_rows(min_row=2), start=1):
            grp = export_df.iloc[i - 1]["Modalidad"] if i - 1 < len(export_df) else ""
            fill = PatternFill("solid", fgColor="D6EAF8") if grp != prev_grp else (alt_fill1 if i % 2 == 0 else alt_fill2)
            prev_grp = grp
            for cell in row_cells:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 25)

        # Hoja 2: Datos Comparativos
        comp_table.to_excel(writer, index=False, sheet_name="Datos_Comparativos")
        ws2 = writer.sheets["Datos_Comparativos"]
        for cell in ws2[1]:
            cell.font  = Font(bold=True, size=9, color="1A1A1A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.fill   = PatternFill("solid", fgColor=_header_color(str(cell.value or "")))
        ws2.row_dimensions[1].height = 40
        for i, row_cells in enumerate(ws2.iter_rows(min_row=2), start=1):
            fill = alt_fill1 if i % 2 == 0 else alt_fill2
            for cell in row_cells:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        diff_cols = [i + 1 for i, h in enumerate(comp_table.columns) if h.startswith("Diff%_")]
        for col_idx in diff_cols:
            cl = get_column_letter(col_idx)
            dr = f"{cl}2:{cl}{len(comp_table) + 1}"
            ws2.conditional_formatting.add(dr, CellIsRule("between",  ["-10", "10"],  fill=PatternFill("solid", fgColor="ABEBC6")))
            ws2.conditional_formatting.add(dr, CellIsRule("between",  ["10",  "20"],  fill=PatternFill("solid", fgColor="F9E79F")))
            ws2.conditional_formatting.add(dr, CellIsRule("between",  ["-20", "-10"], fill=PatternFill("solid", fgColor="F9E79F")))
            ws2.conditional_formatting.add(dr, CellIsRule("greaterThan", ["20"],      fill=PatternFill("solid", fgColor="F1948A")))
            ws2.conditional_formatting.add(dr, CellIsRule("lessThan",    ["-20"],     fill=PatternFill("solid", fgColor="F1948A")))
        ws2.column_dimensions["A"].width = 16
        for col_cells in ws2.columns:
            if col_cells[0].column_letter != "A":
                ws2.column_dimensions[col_cells[0].column_letter].width = 13
        ws2.freeze_panes = "B2"

    log.info(f"  OK {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Reporte HTML
# ─────────────────────────────────────────────────────────────────────────────

def build_html(summary, comp_table, all_fig_paths, out_sum_suv, out_sum_dix, out_path):
    """Assemble and write the interactive HTML report (embedded figures, metrics table,
    heatmaps) to *out_path*."""
    def fig_to_b64(path):
        """Base64-encode the image file at *path* for inline embedding in the HTML report."""
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()

    def color_metric(v):
        """Inline HTML cell style tiering *v* into good/moderate/poor color bands."""
        try:
            v = float(v)
            if v >= 0.90: return "style='background:#d5f5e3;font-weight:bold'"
            if v >= 0.75: return "style='background:#fef9e7'"
            return "style='background:#fdecea'"
        except Exception:
            return ""

    cols    = ["Variable", "Unidades", "n", "r (Pearson)", "rho (Spearman)",
               "CCC", "CCC 95% CI", "ICC (2,1)", "ICC 95% CI",
               "Bias", "LoA inf", "LoA sup", "RMSE", "MAE"]
    headers = ["Variable", "Unidades", "n", "r", "&rho;", "CCC",
               "CCC IC95%", "ICC(2,1)", "ICC IC95%", "Bias", "LoA inf", "LoA sup", "RMSE", "MAE"]
    mrows = ""
    prev_g = None
    for _, row in summary.iterrows():
        if row["Grupo"] != prev_g:
            grp_lbl = row["Grupo"].replace("_", " \u2013 ")
            bg = "#1a3a5c" if row["Modalidad"] == "SUV" else "#1a4a3c"
            mrows += (f"<tr><td colspan='{len(cols)}' style='background:{bg};"
                      f"color:white;font-weight:bold;padding:8px 12px'>{grp_lbl}</td></tr>")
            prev_g = row["Grupo"]
        cells = ""
        for col, hdr in zip(cols, headers):
            val = row.get(col, "")
            raw = val
            if isinstance(val, float) and not np.isnan(val):
                val = f"{val:.3f}"
            elif isinstance(val, float):
                val = "&mdash;"
            style = color_metric(raw) if col in ("CCC", "ICC (2,1)") else ""
            cells += f"<td {style}>{val}</td>"
        mrows += f"<tr>{cells}</tr>"
    mhdr = "".join(f"<th>{h}</th>" for h in headers)

    # Tabla comparativa (preview 30 filas)
    ct = comp_table.head(30)
    show_cols = ["pesa_id"] + [c for c in ct.columns if c.startswith(("Man_", "Auto_", "Diff%_"))]
    chdr  = "".join(f"<th>{c}</th>" for c in show_cols)
    cbody = ""
    for i, (_, row) in enumerate(ct[show_cols].iterrows()):
        bg = "#f8f9fa" if i % 2 == 0 else "white"
        cells = ""
        for col in show_cols:
            val   = row[col]
            style = f"background:{bg}"
            if col.startswith("Diff%_") and pd.notna(val):
                v = float(val)
                style = "background:#d5f5e3" if abs(v) <= 10 else ("background:#fef9e7" if abs(v) <= 20 else "background:#fdecea")
                val = f"{v:+.1f}%"
            elif isinstance(val, float):
                val = f"{val:.3f}" if pd.notna(val) else "&mdash;"
            cells += f"<td style='{style};padding:4px 8px;font-size:12px'>{val}</td>"
        cbody += f"<tr>{cells}</tr>"

    # Secciones de figuras
    fig_sections = ""
    for sec_i, (gkey, ginfo) in enumerate(COMPARISONS.items(), start=3):
        if gkey not in all_fig_paths:
            continue
        n_small = len(ginfo["pairs"]) <= 2
        fig_sections += f"""
<h2 data-editable="true">{sec_i}. {ginfo['title']}</h2>
<div class="fig-wrap">
  <img src="data:image/png;base64,{fig_to_b64(all_fig_paths[gkey])}"
       alt="{gkey}" style="max-width:{'70%' if n_small else '100%'}">
  <div class="fig-caption" data-editable="true">
    <strong>Figura {sec_i}.</strong>
    Izquierda: scatter (identidad rojo, regresi&oacute;n naranja).
    Derecha: Bland-Altman (bias azul, LoA&plusmn;1.96&middot;SD rojo discontinuo).
  </div>
</div>"""

    n_sec = len(COMPARISONS)
    today = date.today().strftime("%d/%m/%Y")

    HTML = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>PESA Fat &mdash; Manual vs Autom&aacute;tico</title>
<style>
  body{{font-family:'Segoe UI',Arial,sans-serif;margin:0;background:#f4f6f9;color:#222;padding-top:46px}}
  .container{{max-width:1200px;margin:auto;padding:30px}}
  h1{{color:#1a3a5c;border-bottom:3px solid #2e86ab;padding-bottom:10px}}
  h2{{color:#1a3a5c;margin-top:40px;border-left:5px solid #2e86ab;padding-left:12px}}
  h3{{color:#2e4057;margin-top:28px}}
  p,li{{line-height:1.7;font-size:15px}}
  .badge{{display:inline-block;padding:2px 10px;border-radius:12px;font-size:12px;font-weight:bold;color:white}}
  .badge-green{{background:#27ae60}} .badge-yellow{{background:#f39c12}} .badge-red{{background:#e74c3c}}
  .warn-box{{background:#fef9e7;border-left:4px solid #f39c12;padding:12px 18px;margin:16px 0;border-radius:4px}}
  .info-box{{background:#eaf4fc;border-left:4px solid #2e86ab;padding:12px 18px;margin:16px 0;border-radius:4px}}
  .metric-def{{background:white;border-radius:8px;padding:16px 22px;margin:10px 0;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
  .metric-def h4{{margin:0 0 6px;color:#1a3a5c;font-size:15px}}
  .metric-def p{{margin:0;font-size:14px;color:#444}}
  table{{border-collapse:collapse;width:100%;background:white;border-radius:8px;overflow:hidden;box-shadow:0 1px 6px rgba(0,0,0,.1);margin:18px 0}}
  th{{background:#2e4057;color:white;padding:10px 12px;font-size:13px;text-align:center}}
  td{{padding:7px 12px;text-align:center;font-size:13px;border-bottom:1px solid #eee}}
  .fig-wrap{{text-align:center;margin:24px 0}}
  .fig-wrap img{{max-width:100%;border-radius:8px;box-shadow:0 2px 10px rgba(0,0,0,.15)}}
  .fig-caption{{font-size:13px;color:#666;margin-top:8px;font-style:italic}}
  footer{{margin-top:50px;text-align:center;font-size:12px;color:#999;border-top:1px solid #ddd;padding-top:16px}}
  #toolbar{{position:fixed;top:0;left:0;right:0;z-index:9999;background:#1a3a5c;color:white;
            display:flex;align-items:center;gap:10px;padding:8px 20px;box-shadow:0 2px 8px rgba(0,0,0,.3)}}
  #toolbar span{{font-weight:bold;font-size:14px;flex:1}}
  .tb-btn{{padding:6px 16px;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:bold}}
  #btn-edit{{background:#f39c12;color:white}} #btn-save{{background:#27ae60;color:white}} #btn-print{{background:#2e86ab;color:white}}
  body.editing [data-editable]{{outline:2px dashed #f39c12;outline-offset:2px;border-radius:3px;cursor:text}}
  body.editing [data-editable]:focus{{outline:2px solid #e67e22;background:#fffdf0}}
  #edit-hint{{display:none;font-size:12px;color:#f9e79f;margin-left:8px}}
  body.editing #edit-hint{{display:inline}}
</style>
<script>
function toggleEdit(){{
  const on=document.body.classList.toggle('editing');
  document.querySelectorAll('[data-editable]').forEach(el=>{{el.contentEditable=on?'true':'false';}});
  const b=document.getElementById('btn-edit');
  b.textContent=on?'Finalizar edici\u00f3n':'Editar texto';
  b.style.background=on?'#e74c3c':'#f39c12';
}}
function saveHTML(){{
  document.body.classList.remove('editing');
  document.querySelectorAll('[data-editable]').forEach(el=>{{el.contentEditable='false';}});
  document.getElementById('btn-edit').textContent='Editar texto';
  document.getElementById('btn-edit').style.background='#f39c12';
  const html='<!DOCTYPE html>\\n'+document.documentElement.outerHTML;
  const blob=new Blob([html],{{type:'text/html;charset=utf-8'}});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);
  a.download='Report_Manual_vs_Auto_editado.html';
  document.body.appendChild(a);a.click();document.body.removeChild(a);URL.revokeObjectURL(a.href);
}}
function printReport(){{window.print()}}
</script>
</head>
<body>
<div id="toolbar">
  <span>PESA Fat &mdash; Report Manual vs Autom&aacute;tico</span>
  <span id="edit-hint">Haz clic en cualquier texto para editarlo</span>
  <button class="tb-btn" id="btn-edit" onclick="toggleEdit()">Editar texto</button>
  <button class="tb-btn" id="btn-save" onclick="saveHTML()">Guardar HTML</button>
  <button class="tb-btn" id="btn-print" onclick="printReport()">Imprimir / PDF</button>
</div>
<div class="container">

<h1 data-editable="true">Comparaci&oacute;n cuantificaciones Manual vs Autom&aacute;tico<br>
<small style="font-size:16px;color:#666">PESA Fat &mdash; Generado el {today}</small></h1>

<h2 data-editable="true">1. Introducci&oacute;n</h2>
<p data-editable="true">Este informe compara las medidas radiol&oacute;gicas obtenidas por
<strong>an&aacute;lisis manual</strong> frente al <strong>pipeline autom&aacute;tico</strong>
de segmentaci&oacute;n y cuantificaci&oacute;n de imagen (PET-CT y MRI-Dixon).</p>
<ul>
  <li data-editable="true"><strong>Fuente 1 (n&asymp;22&ndash;23):</strong> M&eacute;dula &oacute;sea vertebral L3/L4 (SUV) e h&iacute;gado (Dixon FF, VOL, T2*, R2).</li>
  <li data-editable="true"><strong>Fuente 2 (n=6, semana 1):</strong> Ri&ntilde;&oacute;n, p&aacute;ncreas, cu&aacute;driceps, paravertebral (Dixon FF, T2*, VOL); h&iacute;gado, bazo, p&aacute;ncreas, cu&aacute;driceps, paravertebral (SUVmax y SUVmean).</li>
</ul>
<div class="warn-box"><strong>NOTA 1 &mdash; M&eacute;dula &oacute;sea:</strong> <code>MO_SUVMAX</code> en pipeline autom&aacute;tico = m&aacute;ximo sobre la uni&oacute;n L3&cup;L4. Se usa <code>(L3_SUVMAX+L4_SUVMAX)/2</code>.</div>
<div class="warn-box"><strong>NOTA 2 &mdash; R2 hep&aacute;tico:</strong> Escalas manual (~0.5&ndash;3) y autom&aacute;tica (~0.04&ndash;0.10) incompatibles.</div>

<h2 data-editable="true">2. M&eacute;tricas de concordancia</h2>
<div class="metric-def"><h4>Pearson r</h4><p>Correlaci&oacute;n lineal. No detecta sesgos sistem&aacute;ticos.</p></div>
<div class="metric-def"><h4>Spearman &rho;</h4><p>Correlaci&oacute;n de rangos, m&aacute;s robusta ante valores at&iacute;picos.</p></div>
<div class="metric-def"><h4>CCC (Lin, 1989)</h4><p><code>CCC = 2&middot;Cov(X,Y) / [Var(X)+Var(Y)+(&#956;X&#8722;&#956;Y)&sup2;]</code>. IC 95% via Fisher-Z.</p></div>
<div class="metric-def"><h4>ICC(A,1) &mdash; acuerdo absoluto</h4><p>Two-way random, acuerdo absoluto, una medici&oacute;n. Penaliza sesgos.
  <span class="badge badge-red">&lt;0.50</span> pobre &nbsp;
  <span class="badge badge-yellow">0.50&ndash;0.75</span> moderado &nbsp;
  <span class="badge badge-green">0.75&ndash;0.90</span> bueno &nbsp;
  <span class="badge badge-green" style="background:#1a7a3a">&gt;0.90</span> excelente</p></div>
<div class="metric-def"><h4>Bland-Altman (1986)</h4><p>Bias = media(Manual&minus;Auto). LoA = Bias&plusmn;1.96&middot;SD.</p></div>

<h2 data-editable="true">3. Tabla de m&eacute;tricas &mdash; todas las variables</h2>
<div style="overflow-x:auto"><table><tr>{mhdr}</tr>{mrows}</table></div>

{fig_sections}

<h2 data-editable="true">{n_sec+3}. Resumen gr&aacute;fico SUV</h2>
<div class="fig-wrap">
  <img src="data:image/png;base64,{fig_to_b64(out_sum_suv)}" alt="Resumen SUV">
  <div class="fig-caption">Heatmap de correlaci&oacute;n y concordancia &mdash; variables SUV.</div>
</div>

<h2 data-editable="true">{n_sec+4}. Resumen gr&aacute;fico DIXON</h2>
<div class="fig-wrap">
  <img src="data:image/png;base64,{fig_to_b64(out_sum_dix)}" alt="Resumen DIXON">
  <div class="fig-caption">Heatmap de correlaci&oacute;n y concordancia &mdash; variables DIXON.</div>
</div>

<h2 data-editable="true">{n_sec+5}. Tabla comparativa por participante</h2>
<p>Diferencia relativa: <span class="badge badge-green">verde</span> &le;10% &nbsp;
<span class="badge badge-yellow">amarillo</span> 10&ndash;20% &nbsp;
<span class="badge badge-red">rojo</span> &gt;20%.
Tabla completa en <code>Metricas_Manual_vs_Auto.xlsx</code>.</p>
<div style="overflow-x:auto"><table><tr>{chdr}</tr>{cbody}</table></div>

<footer>Generado autom&aacute;ticamente &mdash; PESA Fat Analysis &mdash; {today}<br>
Scripts: <code>CODE/1_build_manual_table.py</code> &rarr; <code>2_build_combined_table.py</code> &rarr; <code>3_run_analysis.py</code></footer>
</div></body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(HTML)
    log.info(f"  OK {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    log.info("Paso 3 — Ejecutando an\u00e1lisis…")

    if not os.path.exists(COMBINED_TABLE_PATH):
        sys.exit(f"ERROR: No se encuentra {COMBINED_TABLE_PATH}\n"
                 "Ejecuta primero: python 2_build_combined_table.py")

    df = pd.read_excel(COMBINED_TABLE_PATH, index_col=0)
    log.info(f"  Tabla combinada: {df.shape[0]} participantes, {df.shape[1]} variables")

    # ── Calcular métricas ─────────────────────────────────────────────────
    log.info("  Calculando m\u00e9tricas\u2026")
    all_metrics = {}
    for gkey, ginfo in COMPARISONS.items():
        for comp in ginfo["pairs"]:
            key = f"{gkey}|{comp['label'].replace(chr(10), ' ')}"
            m = compute_metrics(df[comp["manual"]], df[comp["auto"]])
            m["group"] = gkey; m["modality"] = ginfo["modality"]
            m["label"] = comp["label"].replace("\n", " "); m["units"] = comp["units"]
            all_metrics[key] = m

    # ── Generar figuras ───────────────────────────────────────────────────
    all_fig_paths = {}
    all_ids = np.array(df.index.astype(str))

    def _draw_group(gkey, ginfo, pairs, with_labels):
        """Render one comparison group's scatter + Bland-Altman figure grid and save it to disk."""
        n_pairs = len(pairs)
        fig, axes = plt.subplots(n_pairs, 2, figsize=(13, n_pairs * 3.8), squeeze=False)
        fig.suptitle(ginfo["title"], fontsize=13, fontweight="bold", y=1.01)
        for i, comp in enumerate(pairs):
            key  = f"{gkey}|{comp['label'].replace(chr(10), ' ')}"
            met  = all_metrics.get(key, _NAN_DICT)
            note = comp.get("note", False)
            xv   = df[comp["manual"]].values.astype(float)
            yv   = df[comp["auto"]].values.astype(float)
            plabels = all_ids if with_labels else None
            scatter_panel(axes[i, 0], xv, yv, comp["label"], comp["units"],
                          met, note, comp.get("xlabel"), comp.get("ylabel"), plabels)
            bland_altman_panel(axes[i, 1], xv, yv, comp["label"], comp["units"],
                               met, note, plabels)
        footnotes = []
        if gkey == "SUV_Medula":
            footnotes.append("\u26a0 [NOTA 1]: MO_SUVMAX en pipeline auto contiene MAX(L3,L4); aqu\u00ed se usa (L3_SUVMAX+L4_SUVMAX)/2.")
        if gkey == "DIXON_Higado":
            footnotes.append("\u26a0 [NOTA 2]: R2 hep\u00e1tico \u2014 escalas manual y autom\u00e1tica incompatibles.")
        if footnotes:
            fig.text(0.01, -0.02, "\n".join(footnotes), fontsize=8, color="#C0392B", va="top")
        fig.tight_layout()
        return fig

    for gkey, ginfo in COMPARISONS.items():
        pairs = ginfo["pairs"]
        if ginfo.get("area_figure"):
            pairs = [p for p in pairs
                     if (~(df[p["manual"]].isna() | df[p["auto"]].isna())).sum() >= 1]
            if not pairs:
                log.warning(f"  SKIP {gkey} (todos los pares n=0)")
                continue

        fig = _draw_group(gkey, ginfo, pairs, with_labels=False)
        out_path = os.path.join(OUT_DIR, f"{ginfo['fig_name']}.png")
        fig.savefig(out_path); plt.close(fig)
        all_fig_paths[gkey] = out_path
        log.info(f"  OK {out_path}")

        fig_lbl = _draw_group(gkey, ginfo, pairs, with_labels=True)
        out_lbl = os.path.join(OUT_DIR, f"{ginfo['fig_name']}_labeled.png")
        fig_lbl.savefig(out_lbl); plt.close(fig_lbl)
        log.info(f"  OK {out_lbl}")

    # ── Heatmaps ──────────────────────────────────────────────────────────
    summary_rows = []
    for gkey, ginfo in COMPARISONS.items():
        for comp in ginfo["pairs"]:
            key = f"{gkey}|{comp['label'].replace(chr(10), ' ')}"
            m   = all_metrics[key]
            summary_rows.append({
                "Variable": comp["label"].replace("\n", " "),
                "Grupo": gkey, "Modalidad": ginfo["modality"], "Unidades": comp["units"],
                "n": int(m["n"]) if not np.isnan(m["n"]) else "\u2013",
                "r (Pearson)": m["r"], "rho (Spearman)": m["rho"],
                "CCC": m["ccc"], "CCC 95% CI": f"[{_fmt(m['ccc_lo'])}, {_fmt(m['ccc_hi'])}]",
                "ICC (2,1)": m["icc"], "ICC 95% CI": f"[{_fmt(m['icc_lo'])}, {_fmt(m['icc_hi'])}]",
                "Bias": m["bias"], "LoA inf": m["loa_lo"], "LoA sup": m["loa_hi"],
                "RMSE": m["rmse"], "MAE": m["mae"],
            })
    summary = pd.DataFrame(summary_rows)

    out_sum_suv = os.path.join(OUT_DIR, "Fig_Summary_SUV.png")
    out_sum_dix = os.path.join(OUT_DIR, "Fig_Summary_DIXON.png")
    make_heatmap(summary[summary["Modalidad"] == "SUV"],  "Resumen m\u00e9tricas \u2013 SUV (PET-CT)",  out_sum_suv)
    make_heatmap(summary[summary["Modalidad"] == "DIXON"], "Resumen m\u00e9tricas \u2013 DIXON (MRI)", out_sum_dix)

    # ── Tabla comparativa por participante ─────────────────────────────────
    comp_rows = []
    for pesa_id, row in df.iterrows():
        entry = {"pesa_id": pesa_id}
        for man_col, auto_col, lbl, unit in PAIR_COLS:
            mv = row.get(man_col, np.nan); av = row.get(auto_col, np.nan)
            entry[f"Man_{lbl}"]  = round(float(mv), 4) if pd.notna(mv) else np.nan
            entry[f"Auto_{lbl}"] = round(float(av), 4) if pd.notna(av) else np.nan
            if pd.notna(mv) and pd.notna(av) and float(mv) != 0:
                entry[f"Diff%_{lbl}"] = round(((float(av) - float(mv)) / abs(float(mv))) * 100, 2)
            else:
                entry[f"Diff%_{lbl}"] = np.nan
        comp_rows.append(entry)
    comp_table = pd.DataFrame(comp_rows)
    comp_table = comp_table.dropna(subset=[f"Man_{lbl}" for _, _, lbl, _ in PAIR_COLS], how="all")

    # ── Excel ─────────────────────────────────────────────────────────────
    export_excel(summary, comp_table, METRICS_PATH)

    # ── Consola ───────────────────────────────────────────────────────────
    log.info("\n" + "=" * 95)
    log.info("  RESUMEN DE M\u00c9TRICAS \u2014 MANUAL vs AUTOM\u00c1TICO")
    log.info("=" * 95)
    fmt = "{:<40s} {:>4s}  {:>7s}  {:>7s}  {:>7s}  {:>7s}  {:>8s}  {:>8s}"
    log.info(fmt.format("Variable", "n", "r", "rho", "CCC", "ICC", "Bias", "RMSE"))
    def _f(v):
        """Format *v* to 3 decimals, or ``"-"`` if not a finite float."""
        return f"{v:.3f}" if isinstance(v, float) and not np.isnan(v) else "-"
    prev_grp = None
    for _, row in summary.iterrows():
        if row["Grupo"] != prev_grp:
            log.info("-" * 95)
            log.info(f"  [{row['Grupo']}]")
            prev_grp = row["Grupo"]
        log.info(fmt.format(
            row["Variable"][:40].replace("\u26a0", "!"), str(row["n"]),
            _f(row["r (Pearson)"]), _f(row["rho (Spearman)"]),
            _f(row["CCC"]), _f(row["ICC (2,1)"]),
            _f(row["Bias"]), _f(row["RMSE"]),
        ))
    log.info("=" * 95)

    # ── HTML ──────────────────────────────────────────────────────────────
    build_html(summary, comp_table, all_fig_paths, out_sum_suv, out_sum_dix, REPORT_PATH)

    log.info(f"\nFicheros generados en: {OUT_DIR}")
