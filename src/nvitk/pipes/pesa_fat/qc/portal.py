"""PESA-Fat QC portal (static HTML + Excel and DB-backed review endpoint)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Literal

import json

from nvitk.core.exceptions import BackendUnavailableError
from nvitk.core.logger import Logger

log = Logger()

QcStatus = Literal["PENDING", "OK", "FAIL"]


def review_widget_html(
    *,
    batch: str,
    subject: str,
    pipeline: str,
    structures: Iterable[str],
    report_relpath: str,
) -> str:
    """Self-contained HTML/JS review widget (per-structure QC status/comment rows, Excel
    autosave, DB sync button) embedded into a subject's QC report."""
    from nvitk.pipes.pesa_fat.qc.review_policy import REVIEW_ASPECTS, REVIEW_ASPECT_LABELS

    structs = [str(s).strip() for s in structures if str(s).strip()]
    if not structs:
        return "<p><em>No QC structures defined.</em></p>"
    dom = f"qc_review_{_safe(batch)}_{_safe(subject)}_{_safe(pipeline)}"
    payload = {
        "batch": batch,
        "subject": subject,
        "pipeline": pipeline,
        "report_relpath": report_relpath,
        "structures": structs,
        "aspects": list(REVIEW_ASPECTS),
        "aspect_labels": REVIEW_ASPECT_LABELS,
    }
    return f"""
<div class="card" id="{dom}_card">
  <div class="card-h"><h3>Review</h3><div class="muted">saves to reviews.xlsx · Sync Database pushes to NVITK DB</div></div>
  <div class="card-b">
    <div class="muted" style="margin-bottom:8px">Reviewer: <input id="{dom}_reviewer" placeholder="name" style="padding:6px 8px;border-radius:8px;border:1px solid rgba(229,229,229,0.18);background:rgba(0,0,0,0.25);color:#fff;"/></div>
    <div id="{dom}_rows" style="display:grid;grid-template-columns:1fr;gap:8px"></div>
    <div class="muted" id="{dom}_status" style="margin-top:10px"></div>
    <div style="margin-top:12px">
      <button type="button" id="{dom}_sync" style="padding:8px 14px;border-radius:8px;border:1px solid rgba(252,163,17,0.5);background:rgba(252,163,17,0.15);color:#fca311;font-weight:600;cursor:pointer">Sync Database</button>
    </div>
  </div>
</div>
<script>
(() => {{
  const ctx = {json.dumps(payload)};
  const rows = document.getElementById('{dom}_rows');
  const reviewerInput = document.getElementById('{dom}_reviewer');
  const status = document.getElementById('{dom}_status');
  const syncBtn = document.getElementById('{dom}_sync');
  const commentInputs = {{}};
  const inputStyle = 'padding:6px 8px;border-radius:8px;border:1px solid rgba(229,229,229,0.18);background:rgba(0,0,0,0.25);color:#fff;width:100%;box-sizing:border-box;font:inherit;';
  const mk = (tag, attrs={{}}, txt=null) => {{
    const el = document.createElement(tag);
    for (const [k,v] of Object.entries(attrs)) el.setAttribute(k, v);
    if (txt !== null) el.textContent = txt;
    return el;
  }};
  const rowKey = (structure, aspect) => structure + '::' + aspect;
  const postReview = async (structure, review_aspect, qc_status) => {{
    const reviewer = reviewerInput.value || '';
    const key = rowKey(structure, review_aspect);
    const comment = (commentInputs[key] && commentInputs[key].value) || '';
    const aspectLabel = (ctx.aspect_labels && ctx.aspect_labels[review_aspect]) || review_aspect;
    status.textContent = `Saving ${{structure}} (${{aspectLabel}}) → ${{qc_status}}...`;
    const body = {{...ctx, structure, review_aspect, qc_status, reviewer, comment}};
    const res = await fetch('/review', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify(body),
    }});
    if (!res.ok) {{
      status.textContent = `Save failed (${{res.status}}).`;
      return false;
    }}
    status.textContent = `Saved ${{structure}} (${{aspectLabel}}) → ${{qc_status}}` + (comment ? ' (with comment)' : '') + ' (Excel).';
    return true;
  }};
  const loadState = async () => {{
    const q = new URLSearchParams({{
      batch: ctx.batch,
      subject: ctx.subject,
      pipeline: ctx.pipeline,
    }});
    try {{
      const res = await fetch('/review/state?' + q.toString());
      if (!res.ok) return {{}};
      return await res.json();
    }} catch (e) {{
      return {{}};
    }}
  }};
  const addRow = (structure, review_aspect, saved) => {{
    const key = rowKey(structure, review_aspect);
    const aspectLabel = (ctx.aspect_labels && ctx.aspect_labels[review_aspect]) || review_aspect;
    const wrap = mk('div', {{style:'display:flex;flex-direction:column;gap:6px;padding:10px 12px;border:1px solid rgba(229,229,229,0.18);border-radius:10px;background:rgba(0,0,0,0.20)'}});
    const top = mk('div', {{style:'display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap'}});
    top.appendChild(mk('div', {{style:'font-weight:600'}}, structure + ' · ' + aspectLabel));
    const sel = mk('select', {{style:inputStyle + 'width:auto;min-width:110px'}});
    const initial = (saved && saved.qc_status) ? saved.qc_status : 'PENDING';
    for (const opt of ['PENDING','OK','FAIL']) {{
      const o = mk('option', {{value: opt}}, opt);
      if (opt === initial) o.selected = true;
      sel.appendChild(o);
    }}
    const commentEl = document.createElement('textarea');
    commentEl.setAttribute('rows', '2');
    commentEl.setAttribute('placeholder', 'Comment (optional)');
    commentEl.setAttribute('style', inputStyle + 'resize:vertical;min-height:2.4em;');
    if (saved && saved.comment) commentEl.value = saved.comment;
    commentInputs[key] = commentEl;
    const saveRow = () => postReview(structure, review_aspect, sel.value);
    sel.addEventListener('change', saveRow);
    commentEl.addEventListener('blur', saveRow);
    top.appendChild(sel);
    wrap.appendChild(top);
    wrap.appendChild(commentEl);
    rows.appendChild(wrap);
  }};
  syncBtn.addEventListener('click', async () => {{
    syncBtn.disabled = true;
    status.textContent = 'Syncing to database (may take a moment)...';
    try {{
      const res = await fetch('/review/sync-db', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify(ctx),
      }});
      const data = await res.json().catch(() => ({{}}));
      if (!res.ok || !data.ok) {{
        status.textContent = `Database sync failed: ${{data.error || data.db_error || res.status}}`;
        return;
      }}
      const st = data.stats || {{}};
      status.textContent = `Database synced (${{st.synced_structures || 0}} entries, ${{st.updated_measurements || 0}} measurement rows).`;
    }} catch (e) {{
      status.textContent = 'Database sync failed: ' + e;
    }} finally {{
      syncBtn.disabled = false;
    }}
  }});
  (async () => {{
    const state = await loadState();
    if (state.reviewer) reviewerInput.value = state.reviewer;
    for (const s of ctx.structures) {{
      const savedStruct = (state.structures && state.structures[s]) || {{}};
      for (const aspect of (ctx.aspects || [])) {{
        addRow(s, aspect, savedStruct[aspect] || null);
      }}
    }}
  }})();
}})();
</script>
""".strip()


def _safe(s: str) -> str:
    """*s* with any character outside ``[A-Za-z0-9_-]`` replaced by ``_``, for DOM-safe ids."""
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in str(s))


def _reviewable_entry(structure: str, aspect: str, *, pipeline: str) -> bool:
    """Delegate to :func:`nvitk.pipes.pesa_fat.qc.review_policy.is_reviewable_entry`."""
    from nvitk.pipes.pesa_fat.qc.review_policy import is_reviewable_entry

    return is_reviewable_entry(structure, aspect, pipeline=pipeline)


def _utc_now_iso() -> str:
    """Current UTC timestamp as an ISO-8601 string."""
    return datetime.now(timezone.utc).isoformat()


@dataclass(frozen=True)
class ReviewRow:
    """One QC review entry (structure/aspect status + reviewer metadata) for the reviews sheet."""

    batch: str
    subject: str
    pipeline: str
    structure: str
    review_aspect: str
    qc_status: QcStatus
    reviewer: str = ""
    reviewed_at: str = ""
    comment: str = ""
    report_relpath: str = ""


_HEADERS = [
    "batch",
    "subject",
    "pipeline",
    "structure",
    "review_aspect",
    "qc_status",
    "reviewer",
    "reviewed_at",
    "comment",
    "report_relpath",
]


def upsert_review_row_excel(path: Path, row: ReviewRow) -> None:
    """Insert or update *row* in the ``reviews.xlsx`` workbook at *path*, keyed on
    (batch, subject, pipeline, structure, review_aspect)."""
    try:
        import openpyxl
    except Exception as exc:
        raise BackendUnavailableError('openpyxl is required for QC review Excel writes.') from exc

    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(_HEADERS)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "reviews"
        ws.append(_HEADERS)

    # Map header -> column index
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if header != _HEADERS:
        # normalize: rewrite header row if mismatched
        for i, h in enumerate(_HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)

    def key_match(r: int) -> bool:
        """True if worksheet row *r* has the same (batch, subject, pipeline, structure,
        review_aspect) key as *row*."""
        vals = tuple(ws.cell(row=r, column=i).value for i in range(1, 6))
        want = (row.batch, row.subject, row.pipeline, row.structure, row.review_aspect)
        return vals == want

    target_row = None
    for r in range(2, ws.max_row + 1):
        if key_match(r):
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1

    values = [
        row.batch,
        row.subject,
        row.pipeline,
        row.structure,
        row.review_aspect,
        row.qc_status,
        row.reviewer,
        row.reviewed_at or _utc_now_iso(),
        row.comment,
        row.report_relpath,
    ]
    for i, v in enumerate(values, start=1):
        ws.cell(row=target_row, column=i, value=v)

    wb.save(path)


def create_qc_portal_app(
    *,
    qc_root: Path,
    reviews_xlsx: Path,
    results_root: Path | None = None,
    default_batch: str | None = None,
    publish_db: bool = True,
):
    """Return a FastAPI app serving QC HTML and POST /review.

    The app serves:
    - `GET /` dashboard (when `results_root` is provided by the CLI wrapper)
    - `GET /batch/{batch}` convenience redirect to a batch QC index
    - `GET /files/...` static file server rooted at `results_root`
    - `POST /review` upserts reviews.xlsx only (fast)
    - `GET /review/state` returns saved reviews for a report
    - `POST /review/sync-db` batch-publishes a report to NVITK DB + SQLite index
    """
    try:
        from fastapi import FastAPI
        from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
        from fastapi.staticfiles import StaticFiles
    except Exception as exc:
        raise BackendUnavailableError(
            'QC portal requires FastAPI. Install with: pip install "fastapi>=0.110" "uvicorn>=0.27"'
        ) from exc

    app = FastAPI(title="nvitk PESA-Fat QC portal")

    @app.middleware("http")
    async def _no_cache_html(request, call_next):
        """Middleware: disable caching on HTML responses so reviewers always see live state."""
        response = await call_next(request)
        ctype = response.headers.get("content-type", "")
        if "text/html" in ctype or request.url.path.endswith(".html"):
            response.headers["Cache-Control"] = "no-store"
        return response

    qc_root = Path(qc_root).resolve()
    # Keep backward compatibility: if invoked with a single batch qc_root,
    # serve it under /qc so we can also host a dashboard at / later.
    app.mount("/qc", StaticFiles(directory=str(qc_root), html=True), name="qc")

    results_root_eff = Path(results_root).resolve() if results_root is not None else None
    if results_root_eff is not None:
        # Static file server for full RESULTS tree so relative links work.
        app.mount("/files", StaticFiles(directory=str(results_root_eff), html=True), name="files")

        @app.get("/", response_class=HTMLResponse)
        async def dashboard():
            """``GET /``: render the batch-listing dashboard (with a default-batch link if set)."""
            html = _dashboard_html(results_root=results_root_eff, reviews_xlsx=reviews_xlsx)
            if default_batch:
                # Offer a prominent link to the default batch when provided.
                html = html.replace(
                    "<h1>nvitk PESA-Fat QC portal</h1>",
                    "<h1>nvitk PESA-Fat QC portal</h1>"
                    f"<div class='muted' style='margin-top:6px'>Default batch: "
                    f"<a href='/batch/{_esc(default_batch)}'><code>{_esc(default_batch)}</code></a></div>",
                    1,
                )
            return HTMLResponse(html)

        @app.get("/batch/{batch}")
        async def go_batch(batch: str):
            """``GET /batch/{batch}``: redirect to that batch's QC index HTML."""
            batch = str(batch).strip()
            if not batch:
                return RedirectResponse(url="/")
            return RedirectResponse(url=f"/files/{batch}/res_qc/index.html")
    else:
        # No results root context: default to qc_root index.
        @app.get("/")
        async def root_redirect():
            """``GET /`` (no results-root context): redirect to the single mounted QC index."""
            return RedirectResponse(url="/qc/index.html")

    @app.get("/review/state")
    async def get_review_state(
        batch: str,
        subject: str,
        pipeline: str,
    ):
        """``GET /review/state``: saved review rows (status/reviewer/comment per structure/aspect)
        for one batch/subject/pipeline."""
        batch = str(batch).strip()
        subject = str(subject).strip()
        pipeline = str(pipeline).strip()
        rows = _read_reviews_xlsx(reviews_xlsx)
        structures: dict[str, dict[str, dict[str, str]]] = {}
        reviewer = ""
        for r in rows:
            if r.batch != batch or r.subject != subject or r.pipeline != pipeline:
                continue
            if not _reviewable_entry(r.structure, r.review_aspect, pipeline=pipeline):
                continue
            structures.setdefault(r.structure, {})[r.review_aspect] = {
                "qc_status": r.qc_status,
                "reviewer": r.reviewer,
                "reviewed_at": r.reviewed_at,
                "comment": r.comment,
            }
            if r.reviewer and not reviewer:
                reviewer = r.reviewer
        return JSONResponse({"reviewer": reviewer, "structures": structures})

    @app.post("/review")
    async def post_review(payload: dict[str, Any]):
        """``POST /review``: validate and upsert one review row into ``reviews.xlsx``."""
        from nvitk.pipes.pesa_fat.qc.review_policy import DEFAULT_REVIEW_ASPECT

        try:
            review_aspect = str(payload.get("review_aspect", DEFAULT_REVIEW_ASPECT)).strip().upper()
            row = ReviewRow(
                batch=str(payload.get("batch", "")).strip(),
                subject=str(payload.get("subject", "")).strip(),
                pipeline=str(payload.get("pipeline", "")).strip(),
                structure=str(payload.get("structure", "")).strip(),
                review_aspect=review_aspect,
                qc_status=str(payload.get("qc_status", "PENDING")).strip().upper(),
                reviewer=str(payload.get("reviewer", "")).strip(),
                reviewed_at=_utc_now_iso(),
                comment=str(payload.get("comment", "")).strip(),
                report_relpath=str(payload.get("report_relpath", "")).strip(),
            )
            if row.qc_status not in {"PENDING", "OK", "FAIL"}:
                row = ReviewRow(**{**row.__dict__, "qc_status": "PENDING"})
            if not _reviewable_entry(row.structure, row.review_aspect, pipeline=row.pipeline):
                return JSONResponse(
                    {
                        "ok": False,
                        "error": f"Review entry ({row.structure!r}, {row.review_aspect!r}) is not reviewable",
                    },
                    status_code=400,
                )
            upsert_review_row_excel(reviews_xlsx, row)
            return JSONResponse({"ok": True, "excel": True})
        except Exception as exc:
            log.warning("review write failed: %s", exc)
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

    @app.post("/review/sync-db")
    async def sync_review_db(payload: dict[str, Any]):
        """``POST /review/sync-db``: publish this batch/subject/pipeline's reviewed structures
        and measurements to the NVITK DB + SQLite index."""
        batch = str(payload.get("batch", "")).strip()
        subject = str(payload.get("subject", "")).strip()
        pipeline = str(payload.get("pipeline", "")).strip()
        if not batch or not subject or not pipeline:
            return JSONResponse(
                {"ok": False, "error": "batch, subject, and pipeline are required"},
                status_code=400,
            )
        if not publish_db:
            return JSONResponse(
                {"ok": False, "error": "Database publish disabled for this portal instance"},
                status_code=400,
            )
        try:
            from nvitk.pipes.pesa_fat.common.db_publish import try_sync_qc_reviews_for_report

            all_rows = _read_reviews_xlsx(reviews_xlsx)
            row_dicts = [
                {
                    "batch": r.batch,
                    "subject": r.subject,
                    "pipeline": r.pipeline,
                    "structure": r.structure,
                    "review_aspect": r.review_aspect,
                    "qc_status": r.qc_status,
                    "reviewer": r.reviewer,
                    "reviewed_at": r.reviewed_at,
                    "comment": r.comment,
                    "report_relpath": r.report_relpath,
                }
                for r in all_rows
                if _reviewable_entry(r.structure, r.review_aspect, pipeline=r.pipeline)
            ]
            stats, db_error = try_sync_qc_reviews_for_report(
                batch=batch,
                subject=subject,
                pipeline=pipeline,
                rows=row_dicts,
            )
            if db_error:
                return JSONResponse({"ok": False, "db_error": db_error}, status_code=400)
            return JSONResponse({"ok": True, "stats": stats})
        except Exception as exc:
            log.warning("review DB sync failed: %s", exc)
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

    return app


def _read_reviews_xlsx(path: Path) -> list[ReviewRow]:
    """Best-effort read of the shared reviews workbook."""
    try:
        import openpyxl
    except Exception:
        return []
    p = Path(path)
    if not p.exists():
        return []
    try:
        wb = openpyxl.load_workbook(p)
    except Exception:
        return []
    ws = wb.active
    rows: list[ReviewRow] = []
    # Map header to index
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}
    from nvitk.pipes.pesa_fat.qc.review_policy import DEFAULT_REVIEW_ASPECT

    required = {"batch", "subject", "pipeline", "structure", "qc_status"}
    if not required.issubset(idx.keys()):
        return []
    for r in ws.iter_rows(min_row=2, values_only=True):
        try:
            batch = str(r[idx["batch"]] or "").strip()
            subject = str(r[idx["subject"]] or "").strip()
            pipeline = str(r[idx["pipeline"]] or "").strip()
            structure = str(r[idx["structure"]] or "").strip()
            review_aspect = (
                str(r[idx["review_aspect"]] or DEFAULT_REVIEW_ASPECT).strip().upper()
                if "review_aspect" in idx
                else DEFAULT_REVIEW_ASPECT
            )
            qc_status = str(r[idx["qc_status"]] or "PENDING").strip().upper()
            reviewer = str(r[idx.get("reviewer", -1)] or "").strip() if "reviewer" in idx else ""
            reviewed_at = str(r[idx.get("reviewed_at", -1)] or "").strip() if "reviewed_at" in idx else ""
            comment = str(r[idx.get("comment", -1)] or "").strip() if "comment" in idx else ""
            report_relpath = str(r[idx.get("report_relpath", -1)] or "").strip() if "report_relpath" in idx else ""
        except Exception:
            continue
        if not batch or not subject or not pipeline or not structure:
            continue
        if qc_status not in {"PENDING", "OK", "FAIL"}:
            qc_status = "PENDING"
        rows.append(
            ReviewRow(
                batch=batch,
                subject=subject,
                pipeline=pipeline,
                structure=structure,
                review_aspect=review_aspect,
                qc_status=qc_status,
                reviewer=reviewer,
                reviewed_at=reviewed_at,
                comment=comment,
                report_relpath=report_relpath,
            )
        )
    return rows


def _portal_status(rows: list[ReviewRow], *, pipeline: str) -> tuple[str, str]:
    """Return (label, css_class_suffix) for dashboard cells."""
    from nvitk.pipes.pesa_fat.qc.review_policy import portal_display_status

    by_struct: dict[str, dict[str, dict[str, str]]] = {}
    for r in rows:
        by_struct.setdefault(r.structure, {})[r.review_aspect] = {
            "qc_status": r.qc_status,
            "comment": r.comment,
        }
    label, tone = portal_display_status(by_struct, pipeline=pipeline)
    if label == "REVISED":
        return label, f"revised_{tone}"
    return label, tone


def _discover_batches(results_root: Path) -> list[str]:
    """Batch names under *results_root* that have a ``res_qc/index.html`` report."""
    root = Path(results_root)
    if not root.is_dir():
        return []
    batches: list[str] = []
    for p in sorted(root.iterdir()):
        if not p.is_dir() or p.name.startswith("_"):
            continue
        if (p / "res_qc" / "index.html").is_file():
            batches.append(p.name)
    return batches


def _discover_processed_subjects(results_root: Path) -> list[dict[str, str]]:
    """Return rows with links to CT/Dixon reports for any batch."""
    out: list[dict[str, str]] = []
    for batch in _discover_batches(results_root):
        qc_dir = Path(results_root) / batch / "res_qc"
        subj_dir_items = [p for p in qc_dir.iterdir() if p.is_dir()]
        for sd in sorted(subj_dir_items, key=lambda p: p.name):
            subj = sd.name
            ct = sd / f"qc_ctpet_{subj}.html"
            dx = sd / f"qc_dixon_{subj}.html"
            out.append(
                {
                    "batch": batch,
                    "subject": subj,
                    "ct_href": f"/files/{batch}/res_qc/{subj}/{ct.name}" if ct.is_file() else "",
                    "dx_href": f"/files/{batch}/res_qc/{subj}/{dx.name}" if dx.is_file() else "",
                }
            )
    return out


def _esc(s: str) -> str:
    """HTML-escape *s* (``&<>"``)."""
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _dashboard_html(*, results_root: Path, reviews_xlsx: Path) -> str:
    """Render the portal's batch-listing dashboard page (batches, subjects, review status)."""
    batches = _discover_batches(results_root)
    review_rows = _read_reviews_xlsx(reviews_xlsx)

    # Index reviews by (batch, subject, pipeline)
    by_key: dict[tuple[str, str, str], list[ReviewRow]] = {}
    for r in review_rows:
        by_key.setdefault((r.batch, r.subject, r.pipeline), []).append(r)

    processed = _discover_processed_subjects(results_root)
    subjects = sorted({row["subject"] for row in processed})
    subject_datalist = "\n".join(f"<option value='{_esc(s)}'></option>" for s in subjects)

    table_rows: list[str] = []
    for row in processed:
        b = row["batch"]
        s = row["subject"]
        ct_label, ct_css = _portal_status(by_key.get((b, s, "ct-pet-v5"), []), pipeline="ct-pet-v5")
        dx_label, dx_css = _portal_status(by_key.get((b, s, "dixon-v5"), []), pipeline="dixon-v5")
        ct_link = f"<a href='{_esc(row['ct_href'])}'>CT-PET</a>" if row["ct_href"] else "<span class='muted'>—</span>"
        dx_link = f"<a href='{_esc(row['dx_href'])}'>Dixon</a>" if row["dx_href"] else "<span class='muted'>—</span>"
        table_rows.append(
            "<tr data-subject='"
            + _esc(s).lower()
            + "'>"
            f"<td><code>{_esc(b)}</code></td>"
            f"<td><code>{_esc(s)}</code></td>"
            f"<td class='st {ct_css}'>{_esc(ct_label)}</td>"
            f"<td class='st {dx_css}'>{_esc(dx_label)}</td>"
            f"<td>{ct_link} · {dx_link}</td>"
            "</tr>"
        )

    batch_links = "\n".join(
        f"<li><a href='/batch/{_esc(b)}'><code>{_esc(b)}</code></a></li>" for b in batches
    ) or "<li><em>No batches found under RESULTS root.</em></li>"

    rows_html = "\n".join(table_rows) or (
        "<tr><td colspan='5'><em>No processed subjects found yet.</em></td></tr>"
    )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <title>nvitk PESA-Fat QC portal</title>
  <style>
    body {{
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial;
      margin: 0; padding: 24px 20px 60px;
      background: #14213d; color: #fff;
    }}
    .wrap {{ max-width: 1200px; margin: 0 auto; }}
    h1 {{ margin: 0 0 10px 0; font-size: 22px; }}
    .muted {{ color: rgba(229,229,229,0.9); font-size: 12px; }}
    .card {{
      margin: 14px 0;
      border: 1px solid rgba(229,229,229,0.18);
      background: rgba(0,0,0,0.22);
      border-radius: 14px;
      overflow: hidden;
    }}
    .card-h {{
      padding: 12px 14px;
      border-bottom: 1px solid rgba(229,229,229,0.16);
      display: flex; align-items: baseline; justify-content: space-between; gap: 12px;
    }}
    .card-h h2 {{ margin: 0; font-size: 14px; }}
    .card-b {{ padding: 14px; }}
    a {{ color: #fca311; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .table-wrap {{ overflow: auto; border-radius: 10px; border: 1px solid rgba(229,229,229,0.18); background: rgba(0,0,0,0.25); }}
    table {{ border-collapse: collapse; width: 100%; min-width: 900px; font-size: 12px; }}
    th, td {{ padding: 7px 8px; border-bottom: 1px solid rgba(229,229,229,0.14); text-align: left; }}
    th {{ position: sticky; top: 0; background: rgba(20,33,61,0.92); }}
    code {{ color: #fff; }}
    .st {{ font-weight: 700; }}
    .st.revised_ok {{ color: #7CFC9A; }}
    .st.revised_fail {{ color: #FF6B6B; }}
    .st.pending {{ color: #FFE08A; }}
    .search-bar {{
      width: min(420px, 100%);
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid rgba(229,229,229,0.25);
      background: rgba(0,0,0,0.25);
      color: #fff;
      font-size: 13px;
      margin-bottom: 10px;
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>nvitk PESA-Fat QC portal</h1>
    <div class="muted">Serving results root: <code>{_esc(str(results_root))}</code> · Reviews: <code>{_esc(str(reviews_xlsx))}</code></div>

    <div class="card">
      <div class="card-h"><h2>Batches</h2><div class="muted">click to open batch QC index</div></div>
      <div class="card-b"><ul style="margin:0;padding-left:18px">{batch_links}</ul></div>
    </div>

    <div class="card">
      <div class="card-h"><h2>Processed subjects (all batches)</h2><div class="muted">REVISED (green) when all structures reviewed · REVISED (red) when all FAIL</div></div>
      <div class="card-b">
        <input class="search-bar" id="subject-search" type="search" list="subject-list" placeholder="Search subject…" autocomplete="off"/>
        <datalist id="subject-list">{subject_datalist}</datalist>
        <div class="table-wrap">
          <table>
            <thead>
              <tr>
                <th>Batch</th>
                <th>Subject</th>
                <th>CT-PET</th>
                <th>Dixon</th>
                <th>Reports</th>
              </tr>
            </thead>
            <tbody>
              {rows_html}
            </tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
  <script>
  (function() {{
    var input = document.getElementById('subject-search');
    var tbody = document.querySelector('table tbody');
    if (!input || !tbody) return;
    input.addEventListener('input', function() {{
      var q = (input.value || '').trim().toLowerCase();
      tbody.querySelectorAll('tr[data-subject]').forEach(function(tr) {{
        var subj = tr.getAttribute('data-subject') || '';
        tr.style.display = !q || subj.indexOf(q) >= 0 ? '' : 'none';
      }});
    }});
  }})();
  </script>
</body>
</html>
"""


__all__ = [
    "ReviewRow",
    "create_qc_portal_app",
    "review_widget_html",
    "upsert_review_row_excel",
]

